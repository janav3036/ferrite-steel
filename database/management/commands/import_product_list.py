import os
from decimal import Decimal

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from database.models import Product

DEFAULT_PATH = os.path.join(settings.BASE_DIR, 'Item Data.xlsx')
DATA_START_ROW = 2
COL_DESCRIPTION = 1
COL_ITEM_NO = 2
COL_SORTING = 3
COL_PURCHASING_UOM = 14
COL_SALES_UOM = 15

UOM_MAP = {
    'TONS': 'ton',
    'TON': 'ton',
    'KG': 'kg',
    'MTR': 'mtr',
    'NOS': 'nos',
}

CATEGORY_MAP = {
    'MAIN': 'main',
    'ROLLING': 'rolling',
    'JINDAL': 'jindal',
    'OTHERS': 'others',
}


def _to_unit(raw):
    if not raw:
        return 'ton'
    return UOM_MAP.get(str(raw).strip().upper(), 'ton')


def _to_category(raw):
    """Maps the 'Sorting' column to CATEGORY_CHOICES. Returns '' for anything
    outside main/rolling/jindal/others — left blank rather than guessed."""
    if not raw:
        return ''
    return CATEGORY_MAP.get(str(raw).strip().upper(), '')


class Command(BaseCommand):
    help = (
        'Upsert Product rows from the client item master export (Item Description / '
        'Item No. / Sorting / Purchasing UoM / Sales UoM) by item_no. Existing rows get '
        'product_name/unit refreshed each run, and category refreshed only when Sorting '
        'maps confidently to main/rolling/jindal (left untouched for unmapped values '
        'like "Others" rather than blanked out). rate/quantity/is_active set via admin '
        'are preserved (stock quantity from the file is not used — rate and quantity '
        'always start at 0, filled in manually). HSN code is not present in this file '
        'and is left untouched.'
    )

    def add_arguments(self, parser):
        parser.add_argument('--path', default=DEFAULT_PATH, help='Path to the source .xlsx file')
        parser.add_argument('--dry-run', action='store_true', help='Preview counts without writing to the DB')

    def handle(self, *args, **options):
        path = options['path']
        dry_run = options['dry_run']

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb['Sheet1']

        valid_rows = []
        skipped_blank = 0

        for r in range(DATA_START_ROW, ws.max_row + 1):
            item_no = ws.cell(row=r, column=COL_ITEM_NO).value
            description = ws.cell(row=r, column=COL_DESCRIPTION).value

            if item_no is None and description is None:
                continue
            if not item_no or description is None:
                skipped_blank += 1
                continue

            sales_uom = ws.cell(row=r, column=COL_SALES_UOM).value
            purch_uom = ws.cell(row=r, column=COL_PURCHASING_UOM).value
            sorting = ws.cell(row=r, column=COL_SORTING).value

            valid_rows.append({
                'item_no': str(item_no).strip(),
                'product_name': str(description).strip(),
                'unit': _to_unit(sales_uom or purch_uom),
                'category': _to_category(sorting),
            })

        existing_item_nos = set(
            Product.objects.exclude(item_no='').values_list('item_no', flat=True)
        )
        new_count = sum(1 for row in valid_rows if row['item_no'] not in existing_item_nos)
        update_count = len(valid_rows) - new_count

        self.stdout.write(f'Source rows read: {ws.max_row - DATA_START_ROW + 1}')
        self.stdout.write(f'Valid items: {len(valid_rows)}')
        self.stdout.write(f'Skipped (blank item_no/description): {skipped_blank}')
        self.stdout.write(f'New products to create: {new_count}')
        self.stdout.write(f'Existing products to refresh (name/unit/mapped category): {update_count}')

        if dry_run:
            self.stdout.write(self.style.WARNING('Dry run — no changes made.'))
            return

        created = 0
        updated = 0
        with transaction.atomic():
            for row in valid_rows:
                defaults = {
                    'product_name': row['product_name'],
                    'unit': row['unit'],
                }
                if row['category']:
                    # Only refresh category when the file gives a confident
                    # main/rolling/jindal mapping — never overwrite an existing
                    # row's category with blank (e.g. 'Others' in the source).
                    defaults['category'] = row['category']
                obj, was_created = Product.objects.update_or_create(
                    item_no=row['item_no'],
                    defaults=defaults,
                    create_defaults={
                        'product_name': row['product_name'],
                        'unit': row['unit'],
                        'category': row['category'],
                        'rate': Decimal('0'),
                        'quantity': Decimal('0'),
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

        self.stdout.write(self.style.SUCCESS(
            f'\nDone: created {created} new rows, refreshed {updated} existing rows.'
        ))
