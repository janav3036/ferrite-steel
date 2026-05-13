"""
One-time script: import ProductList_updated.xlsx → database.Product

Usage (from project root, venv active):
    python import_products.py

Clears all existing Product rows, then re-imports from the Excel file.
Sheet: Sheet1  |  Header row: 10  |  Data from row: 11
Columns: HSN CODE, SIZE, LENGTH, TYPE, SUB TYPE, GRADE, QUANTITY, RATE, PIECES, LOCATION
"""
import os
import sys
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ferite_steel.settings')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

import openpyxl
from database.models import Product

XLSX_PATH = os.path.join(os.path.dirname(__file__), 'ProductList_updated.xlsx')
SHEET_NAME = 'Sheet1'
HEADER_ROW = 10

TYPE_MAP = {
    'main': 'main',
    'rolling': 'rolling',
    'plate': 'plate',
}

SUB_TYPE_MAP = {
    'angle': 'angle',
    'channel': 'channel',
    'ub': 'ub',
    'uc': 'uc',
    'beam': 'beam',
    'flat': 'flat',
    'red material': 'red_material',
    'tmt': 'tmt',
}

# Sub-types in the file that have no matching model choice
UNKNOWN_SUB_TYPES = set()


def clean(val):
    return str(val).strip() if val is not None else ''


def to_decimal(val, default=0):
    if val is None:
        return default
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, TypeError):
        return default


def to_int(val):
    if val is None:
        return None
    try:
        return int(str(val).replace(',', '').strip())
    except (ValueError, TypeError):
        return None


def map_type(raw):
    return TYPE_MAP.get(clean(raw).lower(), '')


def map_sub_type(raw):
    key = clean(raw).lower()
    if not key:
        return ''
    mapped = SUB_TYPE_MAP.get(key, None)
    if mapped is None:
        UNKNOWN_SUB_TYPES.add(clean(raw))
        return ''
    return mapped


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # Confirm header
    header = [clean(c) for c in next(
        ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True)
    )]
    print(f"Header: {header}")

    print(f"\nClearing {Product.objects.count()} existing products...")
    Product.objects.all().delete()

    products = []
    skipped = 0
    counter = 0

    for row in ws.iter_rows(min_row=HEADER_ROW + 1, max_row=ws.max_row, values_only=True):
        _, size, length, typ, sub, grade, qty, rate, pieces, location = (row[i] for i in range(10))

        size_str = clean(size)
        if not size_str:
            skipped += 1
            continue

        type_val = map_type(typ)
        if not type_val:
            skipped += 1
            continue

        counter += 1
        hsn_code = f'IMP-{counter:04d}'

        products.append(Product(
            hsn_code=hsn_code,
            type=type_val,
            sub_type=map_sub_type(sub),
            size=size_str,
            length=clean(length),
            grade=clean(grade),
            quantity=to_decimal(qty),
            rate=to_decimal(rate),
            pieces=to_int(pieces),
            location=clean(location),
            is_active=True,
        ))

    Product.objects.bulk_create(products)

    print(f"Imported:  {len(products)} products")
    print(f"Skipped:   {skipped} rows (blank size or unknown type)")
    if UNKNOWN_SUB_TYPES:
        print(f"Unknown sub-types (saved with blank sub_type): {sorted(UNKNOWN_SUB_TYPES)}")
    print("Done.")


if __name__ == '__main__':
    main()
