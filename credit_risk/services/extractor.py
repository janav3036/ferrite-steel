import openpyxl
from .matching import find_best_match


def extract_trading_history(file_obj, filename, customer):
    """Returns {'sales': {...}|None, 'purchase': {...}|None, 'company_context': {...}}."""
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    result = {}
    for sheet_key, sheet_name in (('sales', 'Sales'), ('purchase', 'Purchase')):
        if sheet_name not in wb.sheetnames:
            result[sheet_key] = None
            continue
        result[sheet_key] = _extract_sheet(wb[sheet_name], customer)

    result['company_context'] = _compute_company_context(wb)

    purchase = result.get('purchase')
    if purchase:
        ws = wb['Purchase']
        _, _, total_value_col = _classify_columns(ws)
        rank = _compute_rank(ws, total_value_col, purchase['matched_name'])
        purchase['rank'] = rank
        purchase['in_top_20'] = bool(rank and rank <= 20)
        purchase['streak_flags'] = _streak_flags(purchase['years']) if purchase['in_top_20'] else []

    return result


def _compute_company_context(wb):
    grand_sales = grand_purchase = 0
    if 'Sales' in wb.sheetnames:
        ws = wb['Sales']
        _, _, total_value_col = _classify_columns(ws)
        grand_sales = _grand_total(ws, total_value_col)
    if 'Purchase' in wb.sheetnames:
        ws = wb['Purchase']
        _, _, total_value_col = _classify_columns(ws)
        grand_purchase = _grand_total(ws, total_value_col)
    return {
        'grand_total_sales': grand_sales,
        'grand_total_purchase': grand_purchase,
        'holding_stock': grand_purchase > grand_sales,
    }


def _classify_columns(ws):
    """Returns (year_cols, total_count_col, total_value_col).
    The 'Total' marker lives in row2 for Purchase but row1 for Sales —
    check both rows for it, since either might carry it for a given column."""
    row1 = [c.value for c in ws[1]]
    row2 = [c.value for c in ws[2]]

    filled_year_label = [None] * len(row1)
    last = None
    for i, v in enumerate(row1):
        if v is not None:
            last = v
        filled_year_label[i] = last

    total_count_col = total_value_col = None
    year_cols = []
    for i, header in enumerate(row2):
        h2 = (str(header) if header else '').strip().lower()
        h1 = (str(row1[i]) if row1[i] else '').strip().lower()
        combined = h2 or h1
        if combined.startswith('total') and 'count' in combined:
            total_count_col = i
        elif combined.startswith('total') and 'taxable' in combined:
            total_value_col = i
        elif 'count' in h2:
            year_cols.append((i, 'count', filled_year_label[i]))
        elif 'taxable' in h2:
            year_cols.append((i, 'taxable_value', filled_year_label[i]))

    return year_cols, total_count_col, total_value_col


def _row_names(ws):
    """List of (row_index, name) for every data row from row 3 onward."""
    names = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name and not _is_grand_total_row(name):
            names.append((r, str(name)))
    return names


def _extract_sheet(ws, customer):
    year_cols, total_count_col, total_value_col = _classify_columns(ws)
    rows = _row_names(ws)
    if not rows:
        return None

    row_index_by_name = {name: r for r, name in rows}
    match = find_best_match([customer.name, customer.company], [name for _, name in rows])
    if match is None:
        return None

    r = row_index_by_name[match['matched_name']]

    # year_cols alternates (count, taxable_value) column pairs in left-to-right
    # order — pair positionally, not by label text. The source file has a
    # duplicate "23-24" label on the Purchase sheet; grouping by label would
    # silently collapse two distinct periods into one.
    years = []
    seen_labels = {}
    for j in range(0, len(year_cols) - 1, 2):
        count_col, _, year_label = year_cols[j]
        value_col, _, _ = year_cols[j + 1]
        label = year_label or f'period {j // 2 + 1}'
        seen_labels[label] = seen_labels.get(label, 0) + 1
        display_label = label if seen_labels[label] == 1 else f'{label} (duplicate #{seen_labels[label]})'
        years.append({
            'year': display_label,
            'count': ws.cell(row=r, column=count_col + 1).value,
            'taxable_value': ws.cell(row=r, column=value_col + 1).value,
        })

    return {
        'sheet': ws.title,
        'matched_name': match['matched_name'],
        'match_type': match['match_type'],
        'match_score': match.get('score'),
        'alternatives': match.get('alternatives', []),
        'years': years,
        'total_count': ws.cell(row=r, column=total_count_col + 1).value if total_count_col is not None else None,
        'total_taxable_value': ws.cell(row=r, column=total_value_col + 1).value if total_value_col is not None else None,
    }

def _is_grand_total_row(name):
    if not name:
        return False
    n = str(name).strip().lower()
    return n in ('grand total', 'total', 'grand-total')


def _grand_total(ws, total_value_col):
    if total_value_col is None:
        return 0
    # Prefer Tally's own pre-computed Grand Total row if present — more
    # reliable than re-summing, and avoids double-counting if we didn't skip it.
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if _is_grand_total_row(name):
            v = ws.cell(row=r, column=total_value_col + 1).value
            return v if isinstance(v, (int, float)) else 0

    # Fallback: no Grand Total row found — sum every data row ourselves.
    total = 0
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name or _is_grand_total_row(name):
            continue
        v = ws.cell(row=r, column=total_value_col + 1).value
        if isinstance(v, (int, float)):
            total += v
    return total


def _compute_rank(ws, total_value_col, matched_name):
    """1-indexed rank of matched_name among all rows, sorted by total_taxable_value descending."""
    if total_value_col is None:
        return None
    rows = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name or _is_grand_total_row(name):
            continue
        v = ws.cell(row=r, column=total_value_col + 1).value
        rows.append((str(name), v if isinstance(v, (int, float)) else 0))
    rows.sort(key=lambda x: -x[1])
    for i, (name, _) in enumerate(rows, start=1):
        if name == matched_name:
            return i
    return None

def _streak_flags(years):
    """years: chronologically sorted list of {'year', 'count', 'taxable_value'}."""
    flags = []
    active = [bool(y.get('count')) for y in years]
    if not any(active):
        return flags

    last_active_idx = max(i for i, a in enumerate(active) if a)
    if last_active_idx < len(active) - 1:
        flags.append({
            'type': 'trailing_falloff',
            'weight': 2,
            'detail': f"Active through {years[last_active_idx]['year']}, then no activity in "
                      f"{', '.join(y['year'] for y in years[last_active_idx + 1:])}.",
        })

    for i in range(1, len(active) - 1):
        if not active[i] and any(active[:i]) and any(active[i + 1:]):
            flags.append({
                'type': 'mid_gap',
                'weight': 1,
                'detail': f"No activity in {years[i]['year']} despite activity before and after.",
            })

    return flags