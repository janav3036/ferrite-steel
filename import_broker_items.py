"""
One-time script: import 'Broker list & Item list.xlsx' → database.Broker + database.Product

Usage (from project root, venv active):
    python import_broker_items.py

Brokers: upsert on name (update_or_create).
Products: get_or_create on (category, sub_type, size, make) — skips exact duplicates.
Groups skipped: PIPES, BARS, MS RAIL, WIRE, SCRAP, MIX ITEM (no matching model sub_type).
"""
import os
import sys

import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ferite_steel.settings')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

import openpyxl
from database.models import Broker, Product

XLSX_PATH = os.path.join(os.path.dirname(__file__), 'extra_data', 'Borker list & Item list.xlsx')

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean(val):
    return str(val).strip() if val is not None else ''


# Make column value → (category, make_choice)
MAKE_TO_CATEGORY = {
    'JINDAL':  ('main',    'Jindal'),
    'MAIN':    ('main',    ''),
    'ROLLING': ('rolling', ''),
}

# Item Description first word → sub_type (Structure group only)
DESC_TO_SUB_TYPE = {
    'ANGLE':   'angle',
    'CHANNEL': 'channel',
    'UN':      'channel',   # UN = parallel flange channel
    'UB':      'ub',
    'UC':      'uc',
    'WPB/UC':  'uc',
    'BEAM':    'beam',
    'HBEAM':   'beam',
    'NPB':     'beam',
    'WPB':     'beam',
    'WPB/':    'beam',
}

SKIPPABLE_GROUPS = {'MIX ITEM'}

# Item group → sub_type for non-Structure groups
GROUP_TO_SUB_TYPE = {
    'PLATE':   ('plate', ''),
    'TMT':     ('main',  'tmt'),
    'MS FLAT': ('main',  'flat'),
    'PIPES':   ('main',  'pipe'),
    'BARS':    ('main',  'billet'),
    'MS RAIL': ('main',  'rail'),
    'WIRE':    ('main',  'wire'),
    'SCRAP':   ('main',  'scrap'),
}


def map_item(row):
    """
    row indices (0-based): 2=description, 3=item_no, 4=make, 5=foreign_name, 6=manufacturer, 7=group
    Returns a dict of Product field kwargs, or None to skip.
    """
    description = clean(row[2])
    make_raw    = clean(row[4]).upper()
    size        = clean(row[5])
    group       = clean(row[7])

    if not description or not group:
        return None
    if group in SKIPPABLE_GROUPS:
        return None

    category, make = MAKE_TO_CATEGORY.get(make_raw, ('main', ''))

    if group == 'Structure':
        first_word = description.split()[0].upper()
        sub_type = DESC_TO_SUB_TYPE.get(first_word)
        if sub_type is None:
            return None  # T-sections, SS, FABRICATED, etc. — no matching choice
        if make_raw == 'ROLLING':
            if sub_type not in ('angle', 'channel', 'beam', 'flat'):
                category = 'main'
    elif group == 'MS FLAT':
        sub_type = 'flat'
        category = 'rolling' if make_raw == 'ROLLING' else 'main'
    elif group in GROUP_TO_SUB_TYPE:
        category, sub_type = GROUP_TO_SUB_TYPE[group]
    else:
        return None

    if not size:
        size = description  # fallback to full description if foreign_name is blank

    return dict(
        category=category,
        sub_type=sub_type,
        make=make,
        size=size,
        quantity=0,
        rate=0,
        is_active=True,
    )


# ---------------------------------------------------------------------------
# Brokers
# ---------------------------------------------------------------------------

def import_brokers(ws):
    created = updated = skipped = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        name = clean(row[2])
        if not name:
            continue

        bp_code  = clean(row[1])
        phone    = clean(row[3])
        email    = clean(row[4])
        pan      = clean(row[5])
        gst      = clean(row[7])
        address  = ' '.join(filter(None, [clean(row[8]), clean(row[9])])).strip()
        pincode  = clean(row[10])
        city     = clean(row[11])

        location = ', '.join(filter(None, [city, pincode]))

        notes_parts = []
        if bp_code:  notes_parts.append(f'BP Code: {bp_code}')
        if pan:      notes_parts.append(f'PAN: {pan}')
        if gst:      notes_parts.append(f'GST: {gst}')
        if address:  notes_parts.append(f'Address: {address}')
        notes = '\n'.join(notes_parts)

        broker, was_created = Broker.objects.update_or_create(
            name=name,
            defaults=dict(
                phone=phone[:30],
                email=email,
                location=location,
                notes=notes,
                is_active=True,
            ),
        )

        if was_created:
            created += 1
        else:
            updated += 1

    return created, updated


# ---------------------------------------------------------------------------
# Products
# ---------------------------------------------------------------------------

def import_products(ws):
    created = skipped_dup = skipped_no_map = 0

    for row in ws.iter_rows(min_row=7, values_only=True):
        if not any(row):
            continue

        kwargs = map_item(row)
        if kwargs is None:
            skipped_no_map += 1
            continue

        lookup = {k: kwargs[k] for k in ('category', 'sub_type', 'size', 'make')}
        defaults = {k: kwargs[k] for k in kwargs if k not in lookup}

        if Product.objects.filter(**lookup).exists():
            skipped_dup += 1
        else:
            Product.objects.create(**lookup, **defaults)
            created += 1

    return created, skipped_dup, skipped_no_map


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f'Loading: {XLSX_PATH}')
    wb = openpyxl.load_workbook(XLSX_PATH)

    print('\n── Brokers ──')
    broker_ws = wb['Broker list']
    b_created, b_updated = import_brokers(broker_ws)
    print(f'  Created : {b_created}')
    print(f'  Updated : {b_updated}')
    print(f'  Total   : {Broker.objects.count()} brokers in DB')

    print('\n── Products ──')
    item_ws = wb['Item list']
    p_created, p_dup, p_skip = import_products(item_ws)
    print(f'  Created        : {p_created}')
    print(f'  Skipped (dup)  : {p_dup}')
    print(f'  Skipped (group): {p_skip}  ← MIX ITEM (too vague to categorise)')
    print(f'  Total          : {Product.objects.count()} products in DB')

    print('\nDone.')


if __name__ == '__main__':
    main()
